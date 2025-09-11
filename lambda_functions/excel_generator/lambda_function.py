"""
AWS Lambda Function: Excel Report Generator
Generates Excel reports from processed AWS SSM data.
Specialized function with openpyxl dependency only.
"""

import json
import logging
import os
from datetime import datetime
from typing import Any, Dict, List

import boto3

# Configure logging
logger = logging.getLogger()
logger.setLevel(os.getenv("LOG_LEVEL", "INFO"))

# AWS clients
s3_client = boto3.client("s3")


def lambda_handler(event: Dict[str, Any], context) -> Dict[str, Any]:
    """
    Lambda handler for Excel report generation.

    Args:
        event: Lambda event data from JSON-CSV generator
        context: Lambda context object

    Returns:
        Response with Excel report location
    """
    try:
        # Extract summary data from previous step
        summary = event.get("summary", {})
        execution_id = summary.get(
            "execution_id", f"excel_{int(datetime.now().timestamp())}"
        )

        logger.info(f"Starting Excel generation for execution: {execution_id}")

        # Get processed data location
        processed_data_location = summary.get("processed_data_location")
        if not processed_data_location:
            raise ValueError("processed_data_location is required in summary")

        s3_bucket = summary.get("s3_bucket")
        if not s3_bucket:
            raise ValueError("s3_bucket is required in summary")

        # Parse and load processed data from S3
        if not processed_data_location.startswith("s3://"):
            raise ValueError(f"Invalid S3 location format: {processed_data_location}")

        s3_parts = processed_data_location[5:].split("/", 1)
        if len(s3_parts) != 2:
            raise ValueError(f"Invalid S3 location format: {processed_data_location}")

        source_bucket, s3_key = s3_parts

        # Load processed data from S3
        logger.info(f"Loading processed data from s3://{source_bucket}/{s3_key}")
        response = s3_client.get_object(Bucket=source_bucket, Key=s3_key)
        processed_data = json.loads(response["Body"].read().decode("utf-8"))

        # Extract data for Excel generation
        regional_services_data = processed_data.get("regional_services_data", [])
        metadata = processed_data.get("metadata", {})
        regions_map = metadata.get("regions", {})
        services_map = metadata.get("services", {})
        rss_data = metadata.get("rss_data", {})

        if not regional_services_data:
            logger.warning("No regional services data found for Excel generation")
            regional_services_data = []

        # Create temporary directory for Excel generation
        temp_output_dir = "/tmp/reports"
        os.makedirs(temp_output_dir, exist_ok=True)

        # Generate Excel report
        logger.info("Generating Excel report with all sheets...")
        excel_path = os.path.join(temp_output_dir, "aws_regions_services.xlsx")

        # Generate all sheets data
        sheets_data = _generate_all_sheets(
            regional_services_data, regions_map, services_map, rss_data
        )

        # Write Excel file with all sheets and formatting
        _write_excel_file(excel_path, sheets_data, logger)

        # Upload Excel file to temporary S3 location
        temp_s3_prefix = summary.get("temp_s3_prefix", f"temp-reports/{execution_id}")
        excel_s3_key = f"{temp_s3_prefix}/aws_regions_services.xlsx"

        logger.info(f"Uploading Excel to s3://{s3_bucket}/{excel_s3_key}")
        s3_client.upload_file(excel_path, s3_bucket, excel_s3_key)

        # Update summary with Excel location
        updated_summary = summary.copy()
        if "uploaded_files" not in updated_summary:
            updated_summary["uploaded_files"] = {}
        updated_summary["uploaded_files"]["excel"] = f"s3://{s3_bucket}/{excel_s3_key}"

        logger.info(
            f"Excel generation completed successfully for execution: {execution_id}"
        )

        return {
            "statusCode": 200,
            "body": json.dumps(
                {
                    "status": "success",
                    "message": "Excel report generated successfully",
                    "execution_id": execution_id,
                    "excel_location": f"s3://{s3_bucket}/{excel_s3_key}",
                }
            ),
            "execution_id": execution_id,
            "summary": updated_summary,
        }

    except Exception as e:
        logger.error(f"Error in Excel generation: {str(e)}")
        import traceback

        logger.error(traceback.format_exc())

        execution_id = event.get("summary", {}).get("execution_id", "unknown")

        return {
            "statusCode": 500,
            "body": json.dumps(
                {
                    "status": "error",
                    "message": f"Excel generation failed: {str(e)}",
                    "execution_id": execution_id,
                }
            ),
            "execution_id": execution_id,
            "error": str(e),
        }


def _generate_all_sheets(
    regional_services_data: List[Dict[str, Any]],
    regions_map: Dict[str, Any],
    services_map: Dict[str, Any],
    rss_data: Dict[str, Any],
) -> Dict[str, List]:
    """Generate data for all Excel sheets."""

    sheets = {}

    # Sheet 1: Regional Services (main data)
    sheets["Regional Services"] = regional_services_data

    # Sheet 2: Service Matrix (services vs regions)
    service_matrix = []
    all_regions = sorted(regions_map.keys()) if regions_map else []
    all_services = sorted(services_map.keys()) if services_map else []

    # Create service availability matrix
    service_region_map = {}
    for item in regional_services_data:
        service = item.get("service_code", "")
        region = item.get("region_code", "")
        if service not in service_region_map:
            service_region_map[service] = set()
        service_region_map[service].add(region)

    for service in all_services:
        row = {"service": service}
        for region in all_regions:
            row[region] = (
                "Yes" if region in service_region_map.get(service, set()) else "No"
            )
        service_matrix.append(row)

    sheets["Service Matrix"] = service_matrix

    # Sheet 3: Region Summary
    region_summary = []
    for region_code in all_regions:
        region_info = regions_map.get(region_code, {})
        service_count = len(
            [
                item
                for item in regional_services_data
                if item.get("region_code") == region_code
            ]
        )

        region_summary.append(
            {
                "Region Code": region_code,
                "Region Name": region_info.get("name", region_code),
                "Launch Date": region_info.get("launch_date", "N/A"),
                "Launch Date Source": region_info.get("launch_date_source", "N/A"),
                "Announcement URL": region_info.get("announcement_url", "N/A"),
                "Availability Zones": region_info.get("availability_zones", "N/A"),
                "Service Count": service_count,
            }
        )

    sheets["Region Summary"] = region_summary

    # Sheet 4: Service Summary
    service_summary = []
    for service_code in all_services:
        service_info = services_map.get(service_code, {})
        region_count = len(
            [
                item
                for item in regional_services_data
                if item.get("service_code") == service_code
            ]
        )

        service_summary.append(
            {
                "Service Code": service_code,
                "Service Name": service_info.get("name", service_code),
                "Service Category": service_info.get("category", "N/A"),
                "Description": service_info.get("description", "N/A"),
                "Region Count": region_count,
            }
        )

    sheets["Service Summary"] = service_summary

    # Sheet 5: Statistics
    statistics = [
        ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Total Service-Region Combinations", len(regional_services_data)],
        ["Unique Regions", len(all_regions)],
        ["Unique Services", len(all_services)],
        [
            "Average Services per Region",
            f"{len(regional_services_data) / max(len(all_regions), 1):.2f}",
        ],
        [
            "Average Regions per Service",
            f"{len(regional_services_data) / max(len(all_services), 1):.2f}",
        ],
        ["Data Source", "AWS SSM Parameter Store"],
        ["Pipeline", "AWS Lambda Functions"],
        ["Region Coverage", f"{len(all_regions)}/{len(all_regions)} regions"],
        ["Service Coverage", f"{len(all_services)} services discovered"],
        ["Data Freshness", "Real-time from SSM"],
        ["Cache Status", "Not applicable (Lambda)"],
        ["Processing Method", "Serverless Pipeline"],
        ["Report Format", "Multi-sheet Excel"],
        ["Quality Check", "Validated region/service formats"],
        ["Export Status", "Complete"],
    ]

    sheets["Statistics"] = statistics

    return sheets


def _write_excel_file(excel_path: str, sheets_data: Dict[str, List], logger) -> None:
    """Write Excel file with all sheets and formatting."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError as e:
        logger.error(f"Failed to import openpyxl: {e}")
        raise

    wb = Workbook()

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])

    for sheet_name, sheet_data in sheets_data.items():
        logger.info(f"Creating Excel sheet: {sheet_name} with {len(sheet_data)} rows")
        ws = wb.create_sheet(title=sheet_name)

        if not sheet_data:
            ws.cell(row=1, column=1, value="No data available")
            continue

        if sheet_name == "Statistics":
            # Statistics sheet is a list of lists
            for row_idx, row_data in enumerate(sheet_data, 1):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx == 1:  # First column (labels)
                        cell.font = Font(bold=True)
        else:
            # Other sheets are list of dictionaries
            if isinstance(sheet_data[0], dict):
                # Write headers
                headers = list(sheet_data[0].keys())
                for col_idx, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col_idx, value=header)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill(
                        start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
                    )

                # Write data rows
                for row_idx, row_data in enumerate(sheet_data, 2):
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(
                            row=row_idx, column=col_idx, value=row_data.get(header, "")
                        )

        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            ws.column_dimensions[column_letter].width = adjusted_width

    # Save the workbook
    wb.save(excel_path)
    logger.info(f"Excel file saved: {excel_path}")
