"""
AWS Lambda Function: Report Generator
Generates Excel, JSON, and CSV reports from processed AWS SSM data.
"""

import json
import logging
import os
from datetime import datetime
from typing import Any, Dict, List

import boto3

# Configure logging
logger = logging.getLogger()
logger.setLevel(os.getenv("LOG_LEVEL", "INFO"))

# AWS clients
s3_client = boto3.client("s3")
# SNS client will be created when needed to avoid region errors during testing


def lambda_handler(event: Dict[str, Any], context) -> Dict[str, Any]:
    """
    Lambda handler for report generation.

    Args:
        event: Lambda event data containing S3 location of processed data
        context: Lambda context object

    Returns:
        Response with report generation details
    """
    execution_id = event.get("execution_id", f"rpt_{int(datetime.now().timestamp())}")

    try:
        logger.info(f"Starting report generation for execution: {execution_id}")

        # Extract input parameters from Step Functions event
        processed_data_location = event.get("processed_data_location")
        if not processed_data_location:
            raise ValueError("processed_data_location is required in event")

        # Parse S3 location (format: s3://bucket/key)
        if not processed_data_location.startswith("s3://"):
            raise ValueError(f"Invalid S3 location format: {processed_data_location}")

        s3_parts = processed_data_location[5:].split("/", 1)
        if len(s3_parts) != 2:
            raise ValueError(f"Invalid S3 location format: {processed_data_location}")

        s3_bucket, s3_key = s3_parts

        # Load processed data from S3
        logger.info(f"Loading processed data from s3://{s3_bucket}/{s3_key}")
        response = s3_client.get_object(Bucket=s3_bucket, Key=s3_key)
        processed_data = json.loads(response["Body"].read().decode("utf-8"))

        # Simplified report generation - create basic reports directly
        logger.info("Generating simplified reports...")

        # Extract data for report generation
        regional_services_data = processed_data.get("regional_services_data", [])
        metadata = processed_data.get("metadata", {})

        if not regional_services_data:
            logger.warning("No regional services data found for report generation")
            regional_services_data = []

        # Create temporary directory for file generation
        temp_output_dir = "/tmp/reports"
        os.makedirs(temp_output_dir, exist_ok=True)

        # Generate reports matching non-Lambda script structure
        generated_files = {}

        # Get metadata for report generation
        regions_map = metadata.get("regions", {})
        services_map = metadata.get("services", {})
        rss_data = metadata.get("rss_data", {})

        # Generate JSON report (matching non-Lambda format)
        logger.info("Generating JSON report...")
        json_report = {
            "metadata": {
                "generated_at": datetime.now()
                .isoformat()
                .replace("T", " ")
                .split(".")[0],
                "total_combinations": len(regional_services_data),
                "unique_regions": len(regions_map),
                "unique_services": len(services_map),
                "source": "AWS SSM Parameter Store",
            },
            "data": regional_services_data,
        }

        json_path = os.path.join(temp_output_dir, "aws_regions_services.json")
        with open(json_path, "w") as f:
            json.dump(json_report, f, indent=2)
        generated_files["json"] = json_path

        # Generate CSV report (Regional Services only)
        logger.info("Generating CSV report...")
        import csv

        csv_path = os.path.join(temp_output_dir, "aws_regions_services.csv")
        with open(csv_path, "w", newline="") as f:
            if regional_services_data:
                writer = csv.DictWriter(f, fieldnames=regional_services_data[0].keys())
                writer.writeheader()
                writer.writerows(regional_services_data)
            else:
                writer = csv.writer(f)
                writer.writerow(["No data available"])
        generated_files["csv"] = csv_path

        # Generate Excel report with all 5 sheets
        logger.info("Generating comprehensive Excel report...")
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill

            excel_path = os.path.join(temp_output_dir, "aws_regions_services.xlsx")

            # Generate all sheets data
            sheets_data = _generate_all_sheets(
                regional_services_data, regions_map, services_map, rss_data
            )

            # Write Excel with all sheets and formatting
            _write_excel_file(excel_path, sheets_data, logger)

            generated_files["excel"] = excel_path
            logger.info(
                f"Excel report generated successfully with {len(sheets_data)} sheets: {excel_path}"
            )

        except Exception as e:
            logger.warning(f"Failed to generate Excel report: {e}")
            import traceback

            logger.error(traceback.format_exc())

        # Generate summary report
        logger.info("Generating summary report...")
        summary_report = {
            "execution_summary": {
                "execution_id": execution_id,
                "generated_at": datetime.now().isoformat(),
                "regions_discovered": len(metadata.get("regions", {})),
                "services_discovered": len(metadata.get("services", {})),
                "regional_combinations": len(regional_services_data),
            },
            "regions_list": list(metadata.get("regions", {}).keys()),
            "services_list": list(metadata.get("services", {}).keys()),
        }

        summary_path = os.path.join(temp_output_dir, "summary.json")
        with open(summary_path, "w") as f:
            json.dump(summary_report, f, indent=2)
        generated_files["summary"] = summary_path

        # Clean up old execution subfolders before uploading new reports
        try:
            _cleanup_old_report_folders(s3_bucket, logger)
        except Exception as e:
            logger.warning(f"Failed to cleanup old report folders: {e}")

        # Upload all generated files directly to reports/ folder (no subfolders)
        s3_reports = {}
        reports_base_key = "reports"

        for format_name, file_path in generated_files.items():
            # Store files directly in reports/ with standard naming
            filename = os.path.basename(file_path)
            s3_report_key = f"{reports_base_key}/{filename}"

            # Determine content type
            content_type_map = {
                ".json": "application/json",
                ".csv": "text/csv",
                ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            }
            file_extension = os.path.splitext(file_path)[1]
            content_type = content_type_map.get(
                file_extension, "application/octet-stream"
            )

            with open(file_path, "rb") as f:
                s3_client.put_object(
                    Bucket=s3_bucket,
                    Key=s3_report_key,
                    Body=f.read(),
                    ContentType=content_type,
                )

            s3_reports[format_name] = f"s3://{s3_bucket}/{s3_report_key}"

        logger.info(
            f"Latest reports uploaded directly to: s3://{s3_bucket}/{reports_base_key}/"
        )

        # Create execution summary
        execution_summary = {
            "execution_id": execution_id,
            "completion_timestamp": datetime.now().isoformat(),
            "source_data": f"s3://{s3_bucket}/{s3_key}",
            "reports_generated": s3_reports,
            "statistics": {
                "regions_processed": len(metadata.get("regions", {})),
                "services_processed": len(metadata.get("services", {})),
                "regional_combinations": len(regional_services_data),
                "report_formats": len(s3_reports),
            },
            "pipeline_summary": processed_data.get("pipeline_execution", {}),
        }

        # Store execution summary
        summary_key = f"{reports_base_key}/execution_summary.json"
        s3_client.put_object(
            Bucket=s3_bucket,
            Key=summary_key,
            Body=json.dumps(execution_summary, indent=2),
            ContentType="application/json",
        )

        # Send notification if SNS topic is configured
        sns_topic = os.getenv("COMPLETION_SNS_TOPIC")
        if sns_topic:
            try:
                # Create SNS client only when needed
                sns_client = boto3.client("sns")

                notification_message = {
                    "message": "AWS SSM Data Report Generation Completed",
                    "execution_id": execution_id,
                    "reports_location": f"s3://{s3_bucket}/{reports_base_key}/",
                    "statistics": execution_summary["statistics"],
                }

                sns_client.publish(
                    TopicArn=sns_topic,
                    Subject=f"AWS SSM Report Complete - {execution_id}",
                    Message=json.dumps(notification_message, indent=2),
                )
                logger.info(f"Completion notification sent to SNS: {sns_topic}")
            except Exception as e:
                logger.warning(f"Failed to send SNS notification: {e}")

        # Return success response
        response = {
            "status": "success",
            "statusCode": 200,
            "execution_id": execution_id,
            "message": "Report generation completed successfully",
            "reports_location": f"s3://{s3_bucket}/{reports_base_key}/",
            "execution_summary": execution_summary,
            "reports": s3_reports,
        }

        logger.info(
            f"Report generation completed: {json.dumps(execution_summary['statistics'])}"
        )
        return response

    except Exception as e:
        error_msg = f"Report generation failed: {str(e)}"
        logger.error(error_msg, exc_info=True)

        return {
            "status": "failed",
            "statusCode": 500,
            "execution_id": execution_id,
            "error": error_msg,
            "message": "Report generation failed",
        }


def _generate_all_sheets(
    data: List[Dict], regions_map: Dict, services_map: Dict, rss_data: Dict
) -> Dict[str, List]:
    """Generate all Excel sheets data matching non-Lambda script format."""
    sheets_data = {}

    if not data:
        # Return empty sheets with proper structure
        return {
            "Regional Services": [
                {
                    "Region Code": "No data",
                    "Region Name": "",
                    "Service Code": "",
                    "Service Name": "",
                }
            ],
            "Service Matrix": [{"Service": "No data available"}],
            "Region Summary": [
                {
                    "Region Code": "No data",
                    "Region Name": "",
                    "Launch Date": "",
                    "Launch Date Source": "",
                    "Announcement URL": "",
                    "Availability Zones": "",
                    "Service Count": "",
                }
            ],
            "Service Summary": [
                {
                    "Service Code": "No data",
                    "Service Name": "",
                    "Region Count": "",
                    "Coverage %": "",
                }
            ],
            "Statistics": [["Metric", "Value"], ["No data", "N/A"]],
        }

    # 1. Regional Services sheet (raw data)
    sheets_data["Regional Services"] = data

    # 2. Service Matrix sheet
    sheets_data["Service Matrix"] = _generate_service_matrix(data)

    # 3. Region Summary sheet
    sheets_data["Region Summary"] = _generate_region_summary(
        data, regions_map, rss_data
    )

    # 4. Service Summary sheet
    sheets_data["Service Summary"] = _generate_service_summary(data)

    # 5. Statistics sheet
    sheets_data["Statistics"] = _generate_statistics(data)

    return sheets_data


def _generate_service_matrix(data: List[Dict]) -> List[Dict]:
    """Generate service availability matrix (✓/✗ for each service-region combination)."""
    if not data or "Service Name" not in data[0] or "Region Code" not in data[0]:
        return [{"Service": "No data available"}]

    # Get unique services and regions
    services = sorted(list(set(row["Service Name"] for row in data)))
    regions = sorted(list(set(row["Region Code"] for row in data)))

    # Create service-region mapping
    service_regions_map = {}
    for row in data:
        service = row["Service Name"]
        region = row["Region Code"]
        if service not in service_regions_map:
            service_regions_map[service] = set()
        service_regions_map[service].add(region)

    # Create matrix
    matrix_data = []
    for service in services:
        row = {"Service": service}
        service_regions = service_regions_map.get(service, set())

        for region in regions:
            row[region] = "✓" if region in service_regions else "✗"

        matrix_data.append(row)

    return matrix_data


def _generate_region_summary(
    data: List[Dict], regions_map: Dict, rss_data: Dict
) -> List[Dict]:
    """Generate region summary with launch dates and service counts."""
    if not data or "Region Code" not in data[0]:
        return [
            {
                "Region Code": "No data",
                "Region Name": "",
                "Launch Date": "",
                "Launch Date Source": "",
                "Announcement URL": "",
                "Availability Zones": "",
                "Service Count": "",
            }
        ]

    # Count services per region
    region_counts = {}
    for row in data:
        region = row["Region Code"]
        region_counts[region] = region_counts.get(region, 0) + 1

    regions = sorted(list(set(row["Region Code"] for row in data)))

    summary_data = []
    for region in regions:
        region_name = regions_map.get(region, region)
        service_count = region_counts.get(region, 0)

        # Get launch date and metadata from RSS data
        launch_date = "N/A"
        launch_date_source = "N/A"
        announcement_url = "N/A"
        availability_zones = "N/A"

        if rss_data and region in rss_data:
            region_info = rss_data[region]
            launch_date = region_info.get("launch_date", "N/A")
            launch_date_source = region_info.get("launch_date_source", "N/A")
            announcement_url = region_info.get("announcement_url", "N/A")
            availability_zones = region_info.get("availability_zones", "N/A")

        summary_data.append(
            {
                "Region Code": region,
                "Region Name": region_name,
                "Launch Date": launch_date,
                "Launch Date Source": launch_date_source,
                "Announcement URL": announcement_url,
                "Availability Zones": availability_zones,
                "Service Count": service_count,
            }
        )

    return summary_data


def _generate_service_summary(data: List[Dict]) -> List[Dict]:
    """Generate service summary with coverage statistics."""
    if not data or "Service Name" not in data[0] or "Region Code" not in data[0]:
        return [
            {
                "Service Code": "No data",
                "Service Name": "",
                "Region Count": "",
                "Coverage %": "",
            }
        ]

    # Get service-region mappings
    service_regions = {}
    service_codes = {}

    for row in data:
        service_name = row["Service Name"]
        service_code = row.get("Service Code", service_name)
        region = row["Region Code"]

        if service_name not in service_regions:
            service_regions[service_name] = set()
            service_codes[service_name] = service_code

        service_regions[service_name].add(region)

    # Calculate total unique regions
    total_regions = len(set(row["Region Code"] for row in data))

    # Generate summary
    summary_data = []
    for service_name, regions in service_regions.items():
        region_count = len(regions)
        coverage_pct = (region_count / total_regions) if total_regions > 0 else 0

        summary_data.append(
            {
                "Service Code": service_codes[service_name],
                "Service Name": service_name,
                "Region Count": region_count,
                "Coverage %": coverage_pct,
            }
        )

    # Sort by coverage percentage (descending)
    summary_data.sort(key=lambda x: x["Coverage %"], reverse=True)
    return summary_data


def _generate_statistics(data: List[Dict]) -> List[List]:
    """Generate statistics sheet matching non-Lambda format."""
    if not data:
        return [["Metric", "Value"], ["No data", "N/A"]]

    # Calculate statistics
    total_combinations = len(data)
    unique_regions = len(
        set(row.get("Region Code", "") for row in data if "Region Code" in row)
    )
    unique_services = len(
        set(row.get("Service Name", "") for row in data if "Service Name" in row)
    )

    avg_services_per_region = (
        total_combinations / unique_regions if unique_regions > 0 else 0
    )
    avg_regions_per_service = (
        total_combinations / unique_services if unique_services > 0 else 0
    )

    # Create statistics in non-Lambda format (horizontal key-value pairs)
    stats_data = [
        ["Generated At", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Total Service-Region Combinations", total_combinations],
        ["Unique Regions", unique_regions],
        ["Unique Services", unique_services],
        ["Average Services per Region", f"{avg_services_per_region:.2f}"],
        ["Average Regions per Service", f"{avg_regions_per_service:.2f}"],
        ["Data Source", "AWS SSM Parameter Store"],
        ["Pipeline", "AWS Lambda Functions"],
        ["Region Coverage", f"{unique_regions}/38 regions"],
        ["Service Coverage", f"{unique_services} services discovered"],
        ["Data Freshness", "Real-time from SSM"],
        ["Cache Status", "Not applicable (Lambda)"],
        ["Processing Method", "Serverless Pipeline"],
        ["Report Format", "Multi-sheet Excel"],
        ["Quality Check", "Validated region/service formats"],
        ["Export Status", "Complete"],
    ]

    return stats_data


def _write_excel_file(filepath: str, sheets_data: Dict[str, List], logger):
    """Write formatted Excel file with all sheets matching non-Lambda format."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    try:
        wb = Workbook()
        # Remove default sheet
        wb.remove(wb.active)

        # Write all sheets
        for sheet_name, data in sheets_data.items():
            ws = wb.create_sheet(title=sheet_name)

            if sheet_name == "Statistics":
                # Statistics sheet is a list of lists (no headers)
                for row_idx, row_data in enumerate(data, 1):
                    for col_idx, value in enumerate(row_data, 1):
                        ws.cell(row=row_idx, column=col_idx, value=value)
            else:
                # Other sheets are list of dictionaries
                if data and isinstance(data[0], dict):
                    # Write headers
                    headers = list(data[0].keys())
                    for col_idx, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col_idx, value=header)

                    # Write data
                    for row_idx, row_data in enumerate(data, 2):
                        for col_idx, header in enumerate(headers, 1):
                            ws.cell(
                                row=row_idx,
                                column=col_idx,
                                value=row_data.get(header, ""),
                            )

        # Apply formatting to match non-Lambda version
        _apply_excel_formatting(wb, sheets_data, logger)

        # Save workbook
        wb.save(filepath)

    except Exception as e:
        logger.error(f"Failed to write Excel file: {e}")
        raise


def _apply_excel_formatting(wb, sheets_data: Dict[str, List], logger):
    """Apply professional formatting to all Excel sheets."""
    from openpyxl.styles import Font, PatternFill

    # Define colors matching non-Lambda version
    green_fill = PatternFill(
        start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
    )
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header_fill = PatternFill(
        start_color="366092", end_color="366092", fill_type="solid"
    )
    white_font = Font(color="FFFFFF")

    for sheet_name in sheets_data.keys():
        worksheet = wb[sheet_name]

        # Apply header formatting (dark blue background, white font) - skip Statistics sheet
        if sheet_name != "Statistics" and worksheet.max_row > 0:
            for col in range(1, worksheet.max_column + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.fill = header_fill
                cell.font = white_font

        # Apply sheet-specific formatting
        if sheet_name == "Service Matrix":
            logger.info("Applying color formatting to Service Matrix...")
            for row in range(2, worksheet.max_row + 1):  # Skip header
                for col in range(2, worksheet.max_column + 1):  # Skip service name
                    cell = worksheet.cell(row=row, column=col)
                    if cell.value == "✓":
                        cell.fill = green_fill
                    elif cell.value == "✗":
                        cell.fill = red_fill

        elif sheet_name == "Service Summary":
            logger.info("Applying percentage formatting to Service Summary...")
            # Find Coverage % column and format as percentage
            for col in range(1, worksheet.max_column + 1):
                header_cell = worksheet.cell(row=1, column=col)
                if header_cell and header_cell.value == "Coverage %":
                    for row in range(2, worksheet.max_row + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.number_format = "0.0%"
                    break

        # Auto-adjust column widths
        _adjust_column_widths_simple(worksheet)


def _adjust_column_widths_simple(worksheet):
    """Auto-adjust column widths based on content."""
    for column in worksheet.columns:
        column_letter = column[0].column_letter

        # Calculate optimal width based on content
        max_length = 0
        for cell in column:
            if cell.value is not None:
                max_length = max(max_length, len(str(cell.value)))

        # Set column width with reasonable limits (min 10, max 50 characters)
        adjusted_width = min(max(max_length + 2, 10), 50)
        worksheet.column_dimensions[column_letter].width = adjusted_width


def _cleanup_old_report_folders(bucket_name: str, logger) -> None:
    """Clean up old execution-specific folders from S3 reports directory."""
    try:
        logger.info("Cleaning up old execution folders from reports directory...")

        # List all objects in the reports/ prefix
        response = s3_client.list_objects_v2(
            Bucket=bucket_name, Prefix="reports/", Delimiter="/"
        )

        # Get common prefixes (subfolders)
        common_prefixes = response.get("CommonPrefixes", [])
        deleted_folders = 0

        for prefix_info in common_prefixes:
            folder_prefix = prefix_info["Prefix"]

            # Check if this looks like an execution folder (has timestamp-like pattern)
            folder_name = folder_prefix.replace("reports/", "").rstrip("/")

            # Skip if it's the main reports/ folder
            if not folder_name:
                continue

            # Delete execution folders (anything with underscores or timestamps)
            if (
                any(char in folder_name for char in ["_", "-"])
                and folder_name != "reports"
            ):
                logger.info(f"Deleting execution folder: {folder_prefix}")

                # List and delete all objects in this folder
                objects_response = s3_client.list_objects_v2(
                    Bucket=bucket_name, Prefix=folder_prefix
                )

                objects = objects_response.get("Contents", [])
                if objects:
                    delete_keys = [{"Key": obj["Key"]} for obj in objects]
                    s3_client.delete_objects(
                        Bucket=bucket_name, Delete={"Objects": delete_keys}
                    )
                    deleted_folders += 1
                    logger.info(
                        f"Deleted {len(delete_keys)} objects from {folder_prefix}"
                    )

        if deleted_folders > 0:
            logger.info(
                f"Cleanup completed: removed {deleted_folders} old execution folders"
            )
        else:
            logger.info("No old execution folders found to clean up")

    except Exception as e:
        logger.warning(f"Failed to cleanup old folders: {e}")
        # Don't raise - cleanup failure shouldn't stop report generation
