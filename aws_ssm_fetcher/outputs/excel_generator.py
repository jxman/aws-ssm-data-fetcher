"""Excel output generator for AWS SSM Data Fetcher."""

from typing import Dict, List

import pandas as pd
from openpyxl.styles import Font, NamedStyle, PatternFill

from .base import BaseOutputGenerator, OutputContext, OutputError


class ExcelGenerator(BaseOutputGenerator):
    """Generate comprehensive Excel reports with multiple formatted sheets."""

    def __init__(self, context: OutputContext):
        """Initialize Excel generator.

        Args:
            context: Output context with configuration and metadata
        """
        super().__init__(context)

        # Excel formatting constants
        self.green_fill = PatternFill(
            start_color="C6EFCE", end_color="C6EFCE", fill_type="solid"
        )  # Light green for ✓
        self.red_fill = PatternFill(
            start_color="FFC7CE", end_color="FFC7CE", fill_type="solid"
        )  # Light red for ✗
        self.header_fill = PatternFill(
            start_color="366092", end_color="366092", fill_type="solid"
        )  # Dark blue for headers
        self.white_font = Font(color="FFFFFF")  # White font for headers

    def get_default_filename(self) -> str:
        """Get default filename for Excel output.

        Returns:
            Default Excel filename
        """
        return "aws_regions_services.xlsx"

    def generate(self, data: List[Dict]) -> str:
        """Generate comprehensive Excel file with multiple formatted sheets.

        Args:
            data: List of dictionaries containing regional service data

        Returns:
            Path to generated Excel file

        Raises:
            OutputError: If Excel generation fails
        """
        self.validate_data(data)

        try:
            filename = self.context.filename or self.get_default_filename()
            filepath = self._get_filepath(filename)

            self.logger.info(f"Generating comprehensive Excel report: {filepath}")

            # Generate all sheets using data transformation logic
            sheets_data = self._generate_all_sheets(data)

            # Write to Excel with formatting
            self._write_excel_file(filepath, sheets_data)

            # Log summary
            stats = self._get_data_statistics(data)
            self._log_output_summary(filepath, stats)
            self._log_excel_details(sheets_data, stats)

            return filepath

        except Exception as e:
            raise OutputError(f"Excel generation failed: {e}") from e

    def _generate_all_sheets(self, data: List[Dict]) -> Dict[str, pd.DataFrame]:
        """Generate all Excel sheets data.

        Args:
            data: Regional service data

        Returns:
            Dictionary mapping sheet names to DataFrames
        """
        self.logger.info("Generating Excel sheet data...")

        # Convert to DataFrame for processing
        df = pd.DataFrame(data)
        regions = (
            df["Region Code"].unique().tolist() if "Region Code" in df.columns else []
        )

        # Generate Regional Services sheet (raw data)
        regional_services_df = df.copy()

        # Generate Service Matrix using data transformer
        service_matrix_df = self._generate_service_matrix(data)

        # Generate Region Summary
        region_summary_df = self._generate_region_summary(data, regions)

        # Generate Service Summary
        service_summary_df = self._generate_service_summary(data)

        # Generate Statistics
        statistics_df = self._generate_statistics(data)

        return {
            "Regional Services": regional_services_df,
            "Service Matrix": service_matrix_df,
            "Region Summary": region_summary_df,
            "Service Summary": service_summary_df,
            "Statistics": statistics_df,
        }

    def _generate_service_matrix(self, data: List[Dict]) -> pd.DataFrame:
        """Generate service availability matrix.

        Args:
            data: Regional service data

        Returns:
            DataFrame with services vs regions matrix
        """
        df = pd.DataFrame(data)

        # Get unique services and regions
        services = (
            sorted(df["Service Name"].unique()) if "Service Name" in df.columns else []
        )
        regions = (
            sorted(df["Region Code"].unique()) if "Region Code" in df.columns else []
        )

        if not services or not regions:
            return pd.DataFrame({"Service": [], "Note": ["No data available"]})

        # Create matrix
        matrix_data = []
        for service in services:
            row = {"Service": service}
            service_regions = set(
                df[df["Service Name"] == service]["Region Code"].tolist()
            )

            for region in regions:
                row[region] = "✓" if region in service_regions else "✗"

            matrix_data.append(row)

        return pd.DataFrame(matrix_data)

    def _generate_region_summary(
        self, data: List[Dict], regions: List[str]
    ) -> pd.DataFrame:
        """Generate region summary with statistics.

        Args:
            data: Regional service data
            regions: List of region codes

        Returns:
            DataFrame with region summary
        """
        df = pd.DataFrame(data)
        summary_data = []

        for region in regions:
            region_data = (
                df[df["Region Code"] == region] if "Region Code" in df.columns else df
            )

            # Get region name from context or use code
            region_name = region
            if self.context.region_names and region in self.context.region_names:
                region_name = self.context.region_names[region]

            # Count services in this region
            service_count = len(region_data) if not region_data.empty else 0

            # Get launch date from RSS data if available
            launch_date = "N/A"
            if (
                self.context.rss_data
                and region in self.context.rss_data
                and "launch_date" in self.context.rss_data[region]
            ):
                launch_date = self.context.rss_data[region]["launch_date"]

            summary_data.append(
                {
                    "Region Code": region,
                    "Region Name": region_name,
                    "Service Count": service_count,
                    "Launch Date": launch_date,
                }
            )

        return pd.DataFrame(summary_data)

    def _generate_service_summary(self, data: List[Dict]) -> pd.DataFrame:
        """Generate service summary with coverage statistics.

        Args:
            data: Regional service data

        Returns:
            DataFrame with service summary
        """
        df = pd.DataFrame(data)

        if df.empty or "Service Name" not in df.columns:
            return pd.DataFrame({"Service": [], "Note": ["No data available"]})

        services = df["Service Name"].unique()
        total_regions = (
            df["Region Code"].nunique() if "Region Code" in df.columns else 1
        )
        summary_data = []

        for service in services:
            service_data = df[df["Service Name"] == service]

            # Get service name from context or use discovered name
            display_name = service
            if self.context.service_names and service in self.context.service_names:
                display_name = self.context.service_names[service]

            region_count = (
                service_data["Region Code"].nunique()
                if "Region Code" in service_data.columns
                else 1
            )
            coverage_pct = (region_count / total_regions) if total_regions > 0 else 0

            summary_data.append(
                {
                    "Service Code": service,
                    "Service Name": display_name,
                    "Region Count": region_count,
                    "Coverage %": coverage_pct,
                }
            )

        return pd.DataFrame(summary_data).sort_values("Coverage %", ascending=False)

    def _generate_statistics(self, data: List[Dict]) -> pd.DataFrame:
        """Generate comprehensive statistics.

        Args:
            data: Regional service data

        Returns:
            DataFrame with statistics
        """
        df = pd.DataFrame(data)

        if df.empty:
            return pd.DataFrame({"Metric": ["No data"], "Value": ["N/A"]})

        # Calculate basic statistics
        total_combinations = len(df)
        unique_regions = (
            df["Region Code"].nunique() if "Region Code" in df.columns else 0
        )
        unique_services = (
            df["Service Name"].nunique() if "Service Name" in df.columns else 0
        )

        avg_services_per_region = (
            total_combinations / unique_regions if unique_regions > 0 else 0
        )
        avg_regions_per_service = (
            total_combinations / unique_services if unique_services > 0 else 0
        )

        # Generate statistics data
        stats_data = [
            {
                "Metric": "Total Service-Region Combinations",
                "Value": total_combinations,
            },
            {"Metric": "Unique Regions", "Value": unique_regions},
            {"Metric": "Unique Services", "Value": unique_services},
            {
                "Metric": "Average Services per Region",
                "Value": f"{avg_services_per_region:.2f}",
            },
            {
                "Metric": "Average Regions per Service",
                "Value": f"{avg_regions_per_service:.2f}",
            },
            {"Metric": "Data Source", "Value": "AWS SSM Parameter Store"},
            {
                "Metric": "Generated At",
                "Value": self.context.metadata.get("generated_at", "N/A") if self.context.metadata else "N/A",
            },
        ]

        return pd.DataFrame(stats_data)

    def _write_excel_file(self, filepath: str, sheets_data: Dict[str, pd.DataFrame]):
        """Write formatted Excel file with all sheets.

        Args:
            filepath: Path to Excel file
            sheets_data: Dictionary mapping sheet names to DataFrames
        """
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            # Write all sheets
            for sheet_name, df in sheets_data.items():
                df.to_excel(writer, sheet_name=sheet_name, index=False)

            # Apply formatting to all sheets
            self._apply_excel_formatting(writer, sheets_data)

    def _apply_excel_formatting(
        self, writer: pd.ExcelWriter, sheets_data: Dict[str, pd.DataFrame]
    ):
        """Apply professional formatting to all Excel sheets.

        Args:
            writer: ExcelWriter instance
            sheets_data: Dictionary mapping sheet names to DataFrames
        """
        for sheet_name, df in sheets_data.items():
            worksheet = writer.sheets[sheet_name]

            # Apply header formatting (dark blue background, white font)
            self._format_headers(worksheet)

            # Apply sheet-specific formatting
            if sheet_name == "Service Matrix":
                self._format_service_matrix(worksheet)
            elif sheet_name == "Service Summary":
                self._format_service_summary(worksheet)

            # Auto-adjust column widths
            self._adjust_column_widths(worksheet, df)

    def _format_headers(self, worksheet):
        """Format header row with blue background and white font.

        Args:
            worksheet: openpyxl worksheet
        """
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = self.header_fill
            cell.font = self.white_font

    def _format_service_matrix(self, worksheet):
        """Apply color formatting to service matrix (green for ✓, red for ✗).

        Args:
            worksheet: openpyxl worksheet
        """
        self.logger.info("Applying color formatting to Service Matrix...")

        for row in range(2, worksheet.max_row + 1):  # Skip header
            for col in range(2, worksheet.max_column + 1):  # Skip service name
                cell = worksheet.cell(row=row, column=col)
                if cell.value == "✓":
                    cell.fill = self.green_fill
                elif cell.value == "✗":
                    cell.fill = self.red_fill

    def _format_service_summary(self, worksheet):
        """Apply percentage formatting to coverage column.

        Args:
            worksheet: openpyxl worksheet
        """
        self.logger.info("Applying percentage formatting to Service Summary...")

        # Find the Coverage % column
        coverage_col = None
        for col in range(1, worksheet.max_column + 1):
            header_cell = worksheet.cell(row=1, column=col)
            if header_cell.value == "Coverage %":
                coverage_col = col
                break

        if coverage_col:
            # Create and apply percentage style
            try:
                percentage_style = NamedStyle(name="percentage_custom")
                percentage_style.number_format = "0.0%"

                # Apply to all data cells in Coverage % column
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=coverage_col)
                    cell.style = percentage_style
            except ValueError:
                # Style already exists, apply directly
                for row in range(2, worksheet.max_row + 1):
                    cell = worksheet.cell(row=row, column=coverage_col)
                    cell.number_format = "0.0%"

    def _adjust_column_widths(self, worksheet, df: pd.DataFrame):
        """Auto-adjust column widths based on content.

        Args:
            worksheet: openpyxl worksheet
            df: DataFrame for this sheet
        """
        for column in worksheet.columns:
            column_letter = column[0].column_letter
            column_index = column[0].column - 1

            # Get column name
            column_name = (
                df.columns[column_index] if column_index < len(df.columns) else ""
            )

            # Calculate optimal width
            max_length = len(str(column_name)) if column_name else 0

            # Check data content length (sample first 100 rows for performance)
            if column_index < len(df.columns):
                col_data = df.iloc[: min(100, len(df)), column_index]
                for cell_value in col_data:
                    if cell_value is not None:
                        max_length = max(max_length, len(str(cell_value)))

            # Set column width with reasonable limits (min 10, max 50 characters)
            adjusted_width = min(max(max_length + 2, 10), 50)
            worksheet.column_dimensions[column_letter].width = adjusted_width

    def _log_excel_details(
        self, sheets_data: Dict[str, pd.DataFrame], stats: Dict[str, int]
    ):
        """Log detailed Excel generation summary.

        Args:
            sheets_data: Generated sheets data
            stats: Data statistics
        """
        self.logger.info(f"  - Regional Services: {stats['combinations']} rows")

        if "Service Matrix" in sheets_data:
            matrix_df = sheets_data["Service Matrix"]
            regions_count = len([col for col in matrix_df.columns if col != "Service"])
            self.logger.info(
                f"  - Service Matrix: {len(matrix_df)} services × {regions_count} regions"
            )

        self.logger.info(f"  - Region Summary: {stats['regions']} regions")
        self.logger.info(f"  - Service Summary: {stats['services']} services")

        if "Statistics" in sheets_data:
            self.logger.info(
                f"  - Statistics: {len(sheets_data['Statistics'])} metrics"
            )
