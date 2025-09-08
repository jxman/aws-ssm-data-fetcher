#!/usr/bin/env python3

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font
from datetime import datetime

# Read the existing Excel file
excel_file = 'output/aws_regions_services.xlsx'

try:
    # Load the workbook
    wb = load_workbook(excel_file)
    
    # Read the existing data to get actual counts
    regional_df = pd.read_excel(excel_file, sheet_name='Regional Services')
    
    print(f"Regional Services data: {len(regional_df)} rows")
    print(f"Unique regions: {regional_df['Region Code'].nunique()}")
    print(f"Unique services in Regional data: {regional_df['Service Name'].nunique()}")
    
    # Create the corrected statistics with 395 total services
    stats_data = [
        ['Generator', 'AWS SSM Data Fetcher with Caching v2.0'],
        ['', ''],
        ['Summary Statistics', ''],
        ['Total Regions', regional_df['Region Code'].nunique()],
        ['Total Services', 395],  # Corrected value
        ['Total Combinations', len(regional_df)],
        ['', ''],
        ['Regional Service Distribution', ''],
        ['Avg Services per Region', round(regional_df.groupby('Region Code').size().mean(), 1)],
        ['Max Services (Region)', regional_df.groupby('Region Code').size().max()],
        ['Min Services (Region)', regional_df.groupby('Region Code').size().min()],
        ['', ''],
        ['Service Distribution', ''],
        ['Avg Regions per Service', round(regional_df.groupby('Service Name').size().mean(), 1)],
        ['Max Regions (Service)', regional_df.groupby('Service Name').size().max()],
        ['Min Regions (Service)', regional_df.groupby('Service Name').size().min()]
    ]
    
    # Create DataFrame
    stats_df = pd.DataFrame(stats_data, columns=['Generated At', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
    
    # Remove the old Statistics sheet if it exists
    if 'Statistics' in wb.sheetnames:
        del wb['Statistics']
    
    # Write the corrected statistics to Excel, overwriting existing sheet
    with pd.ExcelWriter(excel_file, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
        stats_df.to_excel(writer, sheet_name='Statistics', index=False)
    
    # Reload the workbook to format the Statistics sheet
    wb = load_workbook(excel_file)
    stats_ws = wb['Statistics']
    
    # Apply header formatting (blue background, white font)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    white_font = Font(color="FFFFFF")
    
    # Format header row
    for cell in stats_ws[1]:
        if cell.value:
            cell.fill = header_fill
            cell.font = white_font
    
    # Auto-adjust column widths
    for col in stats_ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        stats_ws.column_dimensions[column].width = adjusted_width
    
    # Save the workbook
    wb.save(excel_file)
    
    print(f"\n✅ Successfully updated Statistics sheet with 395 total services")
    print(f"File saved: {excel_file}")
    
except Exception as e:
    print(f"❌ Error: {e}")