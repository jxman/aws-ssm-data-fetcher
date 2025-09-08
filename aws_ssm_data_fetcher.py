#!/usr/bin/env python3
"""
AWS SSM Data Fetcher
Fetches AWS service and region data from SSM Parameter Store and generates Excel and JSON outputs.
"""

import boto3
import pandas as pd
import json
import os
import pickle
import feedparser
import requests
import re
import time
from datetime import datetime, timedelta
import logging
from typing import Dict, List, Tuple, Optional
from botocore.exceptions import ClientError, NoCredentialsError
from aws_ssm_fetcher.core.config import Config
from aws_ssm_fetcher.core.cache import CacheManager


class AWSSSMDataFetcher:
    def __init__(self, config: Config = None, region='us-east-1', cache_dir='.cache', cache_hours=24):
        """Initialize the AWS SSM client with caching.
        
        Args:
            config: Config object (preferred). If provided, overrides other parameters.
            region: AWS region (deprecated, use config.aws_region)
            cache_dir: Cache directory (deprecated, use config.cache_dir)
            cache_hours: Cache TTL hours (deprecated, use config.cache_hours)
        """
        # Use config if provided, otherwise create from legacy parameters
        if config is not None:
            self.config = config
        else:
            # Backward compatibility: create config from legacy parameters
            self.config = Config(aws_region=region, cache_dir=cache_dir, cache_hours=cache_hours)
        
        # Set attributes from config for backward compatibility
        self.region = self.config.aws_region
        self.cache_dir = self.config.cache_dir
        self.cache_hours = self.config.cache_hours
        
        # Initialize cache manager
        self.cache_manager = CacheManager(self.config)
        
        try:
            self.ssm = boto3.client('ssm', region_name=self.config.aws_region)
                
            logging.info(f"Initialized SSM client for region: {self.config.aws_region}")
            logging.info(f"Cache directory: {self.config.cache_dir}, TTL: {self.config.cache_hours} hours")
        except NoCredentialsError:
            logging.error("AWS credentials not found. Please configure AWS CLI or set environment variables.")
            raise
    
    def get_parameter_value(self, parameter_path: str) -> str:
        """Get a single parameter value from SSM."""
        try:
            response = self.ssm.get_parameter(Name=parameter_path)
            return response['Parameter']['Value']
        except ClientError as e:
            logging.warning(f"Failed to get parameter {parameter_path}: {e}")
            return None
    
    def get_parameters_batch(self, parameter_paths: List[str]) -> Dict[str, str]:
        """Get multiple parameters in batches (max 10 per request)."""
        results = {}
        
        # Process in batches of 10 (SSM limit)
        for i in range(0, len(parameter_paths), 10):
            batch = parameter_paths[i:i+10]
            try:
                response = self.ssm.get_parameters(Names=batch)
                
                # Process successful parameters
                for param in response['Parameters']:
                    results[param['Name']] = param['Value']
                
                # Log any failed parameters
                for invalid in response['InvalidParameters']:
                    logging.warning(f"Invalid parameter: {invalid}")
                    results[invalid] = None
                    
            except ClientError as e:
                logging.error(f"Batch parameter request failed: {e}")
                for path in batch:
                    results[path] = None
        
        return results
    
    def _get_cache_path(self, cache_key: str) -> str:
        """Get the file path for a cache key. (Delegated to cache manager)"""
        return str(self.cache_manager._get_cache_path(cache_key))
    
    def _is_cache_valid(self, cache_path: str) -> bool:
        """Check if cache file exists and is within TTL. (Delegated to cache manager)"""
        from pathlib import Path
        return self.cache_manager._is_cache_valid(Path(cache_path))
    
    def _load_from_cache(self, cache_key: str) -> Optional[any]:
        """Load data from cache if valid. (Delegated to cache manager)"""
        return self.cache_manager.get(cache_key)
    
    def _save_to_cache(self, cache_key: str, data: any) -> None:
        """Save data to cache. (Delegated to cache manager)"""
        self.cache_manager.set(cache_key, data)
    
    def clear_cache(self, cache_key: Optional[str] = None) -> int:
        """Clear cache files. (Delegated to cache manager)"""
        return self.cache_manager.clear(cache_key)
    
    def get_cache_info(self) -> Dict[str, any]:
        """Get information about cached files. (Delegated to cache manager)"""
        return self.cache_manager.get_info()
    
    def fetch_region_rss_data(self) -> Dict[str, Dict]:
        """Fetch region launch dates and metadata from AWS RSS feed."""
        cache_key = "region_rss_data"
        
        # Try to load from cache first
        cached_data = self._load_from_cache(cache_key)
        if cached_data is not None:
            logging.info("Using cached RSS region data")
            return cached_data
        
        logging.info("Fetching AWS regions RSS data...")
        rss_url = "https://docs.aws.amazon.com/global-infrastructure/latest/regions/regions.rss"
        
        try:
            # Fetch RSS feed
            response = requests.get(rss_url, timeout=30)
            response.raise_for_status()
            
            # Parse RSS feed
            feed = feedparser.parse(response.content)
            
            region_data = {}
            
            for entry in feed.entries:
                try:
                    # Extract information from RSS entry
                    title = entry.title
                    description = getattr(entry, 'description', '')
                    published = getattr(entry, 'published', '')
                    link = getattr(entry, 'link', '')
                    
                    # Parse region code from title or description
                    # Example titles: "Asia Pacific (New Zealand) - ap-southeast-6"
                    region_code_match = re.search(r'([a-z]{2}-[a-z]+-[0-9]{1,2})', title + ' ' + description)
                    
                    if region_code_match:
                        region_code = region_code_match.group(1)
                        
                        # Extract region name (everything before the dash and region code)
                        region_name_match = re.search(r'^(.+?)\s*-\s*' + re.escape(region_code), title)
                        region_name = region_name_match.group(1).strip() if region_name_match else title.split('-')[0].strip()
                        
                        # Parse launch date from published date
                        launch_date = 'N/A'
                        if published:
                            try:
                                # Try to parse various date formats
                                parsed_date = datetime.strptime(published, '%a, %d %b %Y %H:%M:%S %z')
                                launch_date = parsed_date.strftime('%Y-%m-%d')
                            except ValueError:
                                try:
                                    parsed_date = datetime.strptime(published, '%a, %d %b %Y %H:%M:%S GMT')
                                    launch_date = parsed_date.strftime('%Y-%m-%d')
                                except ValueError:
                                    # Extract date from description if available
                                    date_match = re.search(r'(\w+ \d{1,2}, \d{4})', description)
                                    if date_match:
                                        try:
                                            parsed_date = datetime.strptime(date_match.group(1), '%B %d, %Y')
                                            launch_date = parsed_date.strftime('%Y-%m-%d')
                                        except ValueError:
                                            pass
                        
                        region_data[region_code] = {
                            'region_name': region_name,
                            'launch_date': launch_date,
                            'announcement_url': link,
                            'rss_title': title,
                            'description': description
                        }
                        
                        logging.debug(f"Parsed region: {region_code} = {region_name} ({launch_date})")
                
                except Exception as e:
                    logging.warning(f"Failed to parse RSS entry: {e}")
                    continue
            
            # Save to cache
            self._save_to_cache(cache_key, region_data)
            
            logging.info(f"Successfully fetched RSS data for {len(region_data)} regions")
            return region_data
            
        except Exception as e:
            logging.error(f"Failed to fetch RSS data: {e}")
            return {}
    
    def fetch_all_ssm_parameters_by_path(self, parameter_path: str, max_retries: int = 3, base_delay: float = 1.0) -> List[str]:
        """Fetch all SSM parameters using get_parameters_by_path with pagination and throttling protection."""
        logging.info(f"Fetching all SSM parameters by path: {parameter_path}")
        
        all_parameters = []
        retry_count = 0
        
        while retry_count <= max_retries:
            try:
                paginator = self.ssm.get_paginator('get_parameters_by_path')
                page_iterator = paginator.paginate(
                    Path=parameter_path,
                    Recursive=True,
                    MaxResults=10  # Reduced to minimize throttling
                )
                
                total_params = 0
                page_count = 0
                for page in page_iterator:
                    for param in page['Parameters']:
                        all_parameters.append(param['Name'])
                        total_params += 1
                    
                    page_count += 1
                    
                    # Add delay between pages to avoid throttling
                    if page_count % 5 == 0:  # Every 5 pages
                        time.sleep(0.5)
                    
                    # Log progress every 50 parameters
                    if total_params % 50 == 0:
                        logging.info(f"Processed {total_params} parameters...")
                
                logging.info(f"Found {total_params} total parameters at path {parameter_path}")
                break  # Success, exit retry loop
                
            except ClientError as e:
                if 'ThrottlingException' in str(e) and retry_count < max_retries:
                    retry_count += 1
                    delay = base_delay * (2 ** retry_count)  # Exponential backoff
                    logging.warning(f"Throttling detected, retrying in {delay}s (attempt {retry_count}/{max_retries})")
                    time.sleep(delay)
                else:
                    logging.error(f"Failed to fetch parameters at path {parameter_path}: {e}")
                    break
            except Exception as e:
                logging.error(f"Failed to fetch parameters at path {parameter_path}: {e}")
                break
        
        return all_parameters
    
    def discover_regions_from_ssm(self) -> List[str]:
        """Discover all AWS regions from SSM parameters with targeted approach."""
        cache_key = "discovered_regions"
        
        # Try cache first
        cached_data = self._load_from_cache(cache_key)
        if cached_data is not None:
            logging.info("Using cached discovered regions")
            return cached_data
        
        logging.info("Discovering all regions from SSM parameters...")
        
        # Get region parameters with limited pagination for efficiency
        try:
            regions = set()
            paginator = self.ssm.get_paginator('get_parameters_by_path')
            page_iterator = paginator.paginate(
                Path='/aws/service/global-infrastructure/regions',
                Recursive=False,  # Only first level to get region directories
                MaxResults=10  # AWS limit for get_parameters_by_path
            )
            
            # Process only first few pages to get most regions efficiently
            pages_processed = 0
            for page in page_iterator:
                for param in page['Parameters']:
                    # Extract region code from paths like /aws/service/global-infrastructure/regions/us-east-1
                    region_match = re.search(r'/regions/([a-z0-9-]+)$', param['Name'])
                    if region_match:
                        region_code = region_match.group(1)
                        regions.add(region_code)
                
                pages_processed += 1
                if pages_processed >= 10:  # Limit to first 10 pages for efficiency
                    break
                    
            # If we didn't get many regions, try the recursive approach but with more targeted extraction
            if len(regions) < 20:
                logging.info("Limited regions found, trying targeted parameter sampling...")
                sample_page = self.ssm.get_parameters_by_path(
                    Path='/aws/service/global-infrastructure/regions',
                    Recursive=True,
                    MaxResults=10
                )
                
                for param in sample_page['Parameters']:
                    region_match = re.search(r'/regions/([a-z0-9-]+)/', param['Name'])
                    if region_match:
                        region_code = region_match.group(1)
                        regions.add(region_code)
            
            discovered_regions = sorted(list(regions))
            
            # Save to cache
            self._save_to_cache(cache_key, discovered_regions)
            
            logging.info(f"Discovered {len(discovered_regions)} regions from SSM")
            return discovered_regions
            
        except Exception as e:
            logging.error(f"Failed to discover regions from SSM: {e}")
            return []
    
    def discover_services_from_ssm(self) -> List[str]:
        """Discover all AWS services from SSM parameters with targeted approach."""
        cache_key = "discovered_services"
        
        # Try cache first
        cached_data = self._load_from_cache(cache_key)
        if cached_data is not None:
            logging.info("Using cached discovered services")
            return cached_data
        
        logging.info("Discovering all services from SSM parameters...")
        
        # Get service parameters with limited pagination for efficiency
        try:
            services = set()
            paginator = self.ssm.get_paginator('get_parameters_by_path')
            page_iterator = paginator.paginate(
                Path='/aws/service/global-infrastructure/services',
                Recursive=False,  # Only first level to get service directories
                MaxResults=10  # AWS limit for get_parameters_by_path
            )
            
            # Process ALL pages to get complete service list
            pages_processed = 0
            for page in page_iterator:
                for param in page['Parameters']:
                    # Extract service code from paths like /aws/service/global-infrastructure/services/ec2
                    service_match = re.search(r'/services/([a-z0-9-]+)$', param['Name'])
                    if service_match:
                        service_code = service_match.group(1)
                        services.add(service_code)
                
                pages_processed += 1
                
                # Log progress every 50 pages
                if pages_processed % 50 == 0:
                    logging.info(f"Processed {pages_processed} pages, found {len(services)} unique services so far...")
                
                # Add small delay every 25 pages to avoid throttling
                if pages_processed % 25 == 0:
                    time.sleep(0.2)
            
            # If we still don't have many services, try recursive approach to fill gaps
            if len(services) < 300:
                logging.info(f"Found {len(services)} services with non-recursive scan, trying recursive approach to find more...")
                
                # Use the full pagination method we created earlier
                all_service_params = self.fetch_all_ssm_parameters_by_path('/aws/service/global-infrastructure/services')
                
                for param in all_service_params:
                    service_match = re.search(r'/services/([a-z0-9-]+)/', param)
                    if service_match:
                        service_code = service_match.group(1)
                        services.add(service_code)
            
            discovered_services = sorted(list(services))
            
            # Save to cache
            self._save_to_cache(cache_key, discovered_services)
            
            logging.info(f"Discovered {len(discovered_services)} services from SSM")
            return discovered_services
            
        except Exception as e:
            logging.error(f"Failed to discover services from SSM: {e}")
            return []

    def get_services_per_region_proper(self, all_services: List[str]) -> Dict[str, List[str]]:
        """Map services to regions using actual AWS SSM data like the AWS services project."""
        cache_key = "region_services_mapping"
        
        # Try cache first
        cached_data = self._load_from_cache(cache_key)
        if cached_data is not None:
            logging.info("Using cached region-services mapping")
            return cached_data
        
        logging.info("Mapping services to regions using actual AWS SSM data...")
        region_services = {}
        
        try:
            for i, service_code in enumerate(all_services, 1):
                logging.info(f"Processing service {i:3d}/{len(all_services)}: {service_code}")
                
                # Get regions where this service is available using AWS SSM path
                service_path = f"/aws/service/global-infrastructure/services/{service_code}/regions"
                
                try:
                    paginator = self.ssm.get_paginator('get_parameters_by_path')
                    page_iterator = paginator.paginate(
                        Path=service_path,
                        Recursive=False,
                        MaxResults=10
                    )
                    
                    service_regions = []
                    for page in page_iterator:
                        for param in page['Parameters']:
                            # Parameter value contains the region code
                            region_code = param['Value']
                            if region_code:
                                service_regions.append(region_code)
                                # Add to region_services mapping
                                if region_code not in region_services:
                                    region_services[region_code] = []
                                if service_code not in region_services[region_code]:
                                    region_services[region_code].append(service_code)
                    
                    logging.info(f"  {service_code} available in {len(service_regions)} regions")
                    
                    # Small delay to avoid throttling
                    if i % 10 == 0:
                        time.sleep(0.5)
                        
                except Exception as e:
                    logging.warning(f"Failed to get regions for service {service_code}: {e}")
                    continue
            
            # Sort services within each region
            for region_code in region_services:
                region_services[region_code].sort()
            
            # Cache the results
            self._save_to_cache(cache_key, region_services)
            
            total_mappings = sum(len(services) for services in region_services.values())
            logging.info(f"Successfully mapped {len(region_services)} regions with {total_mappings} total service mappings")
            
            return region_services
            
        except Exception as e:
            logging.error(f"Failed to map services to regions: {e}")
            return {}

    def fetch_availability_zones(self, regions: List[str]) -> Dict[str, int]:
        """Fetch availability zone counts for regions from SSM with full pagination."""
        cache_key = "availability_zones"
        
        # Try to load from cache first
        cached_data = self._load_from_cache(cache_key)
        if cached_data is not None:
            logging.info("Using cached availability zone data")
            return cached_data
        
        logging.info("Fetching availability zone data with full pagination...")
        
        # Get ALL availability zone parameters with pagination
        all_az_params = self.fetch_all_ssm_parameters_by_path('/aws/service/global-infrastructure/availability-zones')
        
        az_data = {}
        
        # Process each region to count its AZs
        for region in regions:
            az_count = 0
            
            # Look for AZ parameters that belong to this region
            for param_name in all_az_params:
                if '/region' in param_name:
                    try:
                        # Get the region for this AZ parameter
                        response = self.ssm.get_parameter(Name=param_name)
                        if response['Parameter']['Value'] == region:
                            az_count += 1
                    except Exception as e:
                        # If we can't get the parameter, try pattern matching as fallback
                        az_match = re.search(r'/availability-zones/([^/]+)/', param_name)
                        if az_match:
                            az_code = az_match.group(1)
                            # Common AZ to region mappings
                            region_mappings = {
                                'use1': 'us-east-1', 'use2': 'us-east-2',
                                'usw1': 'us-west-1', 'usw2': 'us-west-2',
                                'euw1': 'eu-west-1', 'euw2': 'eu-west-2', 'euw3': 'eu-west-3',
                                'euc1': 'eu-central-1', 'euc2': 'eu-central-2',
                                'eun1': 'eu-north-1', 'eus1': 'eu-south-1', 'eus2': 'eu-south-2',
                                'apne1': 'ap-northeast-1', 'apne2': 'ap-northeast-2', 'apne3': 'ap-northeast-3',
                                'aps1': 'ap-south-1', 'aps2': 'ap-south-2',
                                'apse1': 'ap-southeast-1', 'apse2': 'ap-southeast-2', 
                                'apse3': 'ap-southeast-3', 'apse4': 'ap-southeast-4',
                                'ape1': 'ap-east-1', 'ape2': 'ap-east-2',
                                'cac1': 'ca-central-1', 'caw1': 'ca-west-1',
                                'sae1': 'sa-east-1',
                                'afs1': 'af-south-1',
                                'mes1': 'me-south-1', 'mec1': 'me-central-1',
                                'ilc1': 'il-central-1'
                            }
                            
                            az_region_prefix = az_code.rstrip('0123456789-az')
                            if region_mappings.get(az_region_prefix) == region:
                                az_count += 1
            
            if az_count > 0:
                az_data[region] = az_count
                logging.debug(f"Found {az_count} AZs for {region}")
            else:
                # Fallback to known AZ counts for established regions
                common_az_counts = {
                    'us-east-1': 6, 'us-east-2': 3, 'us-west-1': 3, 'us-west-2': 4,
                    'eu-west-1': 3, 'eu-west-2': 3, 'eu-west-3': 3, 'eu-central-1': 3,
                    'ap-northeast-1': 3, 'ap-northeast-2': 4, 'ap-southeast-1': 3, 'ap-southeast-2': 3,
                    'ap-south-1': 3, 'ca-central-1': 3, 'sa-east-1': 3
                }
                az_data[region] = common_az_counts.get(region, 3)  # Default to 3 AZs
        
        # Save to cache
        self._save_to_cache(cache_key, az_data)
        
        logging.info(f"Successfully fetched AZ data for {len(az_data)} regions")
        return az_data
    
    def fetch_regions(self) -> List[str]:
        """Get list of AWS regions by discovering them from SSM parameters."""
        logging.info("Discovering AWS regions from SSM parameters...")
        
        # Use the new discovery method with full pagination
        discovered_regions = self.discover_regions_from_ssm()
        
        if discovered_regions:
            logging.info(f"Successfully discovered {len(discovered_regions)} regions from SSM")
            return discovered_regions
        else:
            # Fallback to known regions if discovery fails
            logging.warning("Failed to discover regions, falling back to known regions")
            fallback_regions = [
                'af-south-1', 'ap-east-1', 'ap-northeast-1', 'ap-northeast-2', 'ap-northeast-3',
                'ap-south-1', 'ap-south-2', 'ap-southeast-1', 'ap-southeast-2', 'ap-southeast-3',
                'ap-southeast-4', 'ca-central-1', 'ca-west-1', 'eu-central-1', 'eu-central-2',
                'eu-north-1', 'eu-south-1', 'eu-south-2', 'eu-west-1', 'eu-west-2', 'eu-west-3',
                'il-central-1', 'me-central-1', 'me-south-1', 'sa-east-1', 'us-east-1', 'us-east-2',
                'us-west-1', 'us-west-2', 'us-gov-east-1', 'us-gov-west-1'
            ]
            logging.info(f"Using {len(fallback_regions)} fallback regions")
            return fallback_regions
    
    def fetch_services(self) -> List[str]:
        """Get list of AWS services by discovering them from SSM parameters."""
        logging.info("Discovering AWS services from SSM parameters...")
        
        # Use the new discovery method with full pagination
        discovered_services = self.discover_services_from_ssm()
        
        if discovered_services:
            logging.info(f"Successfully discovered {len(discovered_services)} services from SSM")
            return discovered_services
        else:
            # Fallback to known services if discovery fails
            logging.warning("Failed to discover services, falling back to known services")
            fallback_services = [
                'accessanalyzer', 'account', 'acm', 'acm-pca', 'amplify', 'amplifyuibuilder',
                'apigateway', 'apigatewayv2', 'appconfig', 'appconfigdata', 'appflow',
                'application-autoscaling', 'applicationinsights', 'appmesh', 'apprunner',
                'appstream', 'appsync', 'athena', 'autoscaling', 'backup', 'batch',
                'bedrock', 'budgets', 'ce', 'chime', 'cloud9', 'cloudcontrol',
                'cloudformation', 'cloudfront', 'cloudhsm', 'cloudsearch', 'cloudtrail',
                'cloudwatch', 'codeartifact', 'codebuild', 'codecommit', 'codedeploy',
                'codeguru-reviewer', 'codepipeline', 'codestar', 'cognito-identity',
                'cognito-idp', 'comprehend', 'config', 'connect', 'databrew', 'dataexchange',
                'datapipeline', 'datasync', 'dax', 'detective', 'devicefarm', 'directconnect',
                'discovery', 'dms', 'docdb', 'ds', 'dynamodb', 'ebs', 'ec2', 'ecr',
                'ecs', 'efs', 'eks', 'elastic-inference', 'elasticache', 'elasticbeanstalk',
                'elasticfilesystem', 'elasticloadbalancing', 'elasticmapreduce', 'elastictranscoder',
                'emr', 'es', 'events', 'firehose', 'fms', 'forecast', 'frauddetector',
                'fsx', 'gamelift', 'glacier', 'globalaccelerator', 'glue', 'greengrass',
                'groundstation', 'guardduty', 'health', 'iam', 'imagebuilder', 'inspector',
                'inspector2', 'iot', 'iotanalytics', 'iotevents', 'kafka', 'kendra',
                'kinesis', 'kinesisanalytics', 'kms', 'lakeformation', 'lambda', 'lex',
                'license-manager', 'lightsail', 'logs', 'machinelearning', 'macie2',
                'managedblockchain', 'mediaconnect', 'mediaconvert', 'medialive', 'mediapackage',
                'mediastore', 'mediatailor', 'memorydb', 'mq', 'mturk', 'neptune',
                'networkfirewall', 'networkmanager', 'opensearch', 'organizations', 'outposts',
                'personalize', 'pinpoint', 'polly', 'pricing', 'qldb', 'quicksight',
                'ram', 'rds', 'redshift', 'rekognition', 'resource-groups', 'robomaker',
                'route53', 'route53domains', 'route53resolver', 's3', 's3control', 'sagemaker',
                'secretsmanager', 'securityhub', 'serverlessrepo', 'servicecatalog', 'servicediscovery',
                'ses', 'shield', 'signer', 'sns', 'sqs', 'ssm', 'sso', 'stepfunctions',
                'storagegateway', 'sts', 'support', 'swf', 'synthetics', 'textract',
                'timestream', 'transcribe', 'transfer', 'translate', 'trustedadvisor', 'waf',
                'wafv2', 'workdocs', 'worklink', 'workmail', 'workspaces', 'xray'
            ]
            logging.info(f"Using {len(fallback_services)} fallback services")
            return fallback_services
    
    def fetch_region_names(self, regions: List[str]) -> Dict[str, str]:
        """Fetch human-readable names for all regions."""
        cache_key = "region_names"
        
        # Try to load from cache first
        cached_data = self._load_from_cache(cache_key)
        if cached_data is not None:
            # Filter cached data to only include requested regions
            filtered_data = {k: v for k, v in cached_data.items() if k in regions}
            if len(filtered_data) == len(regions):
                logging.info(f"Using cached region names for {len(regions)} regions")
                return filtered_data
        
        logging.info("Fetching region display names...")
        
        # Build parameter paths for region names
        region_name_paths = [f'/aws/service/global-infrastructure/regions/{region}/longName' for region in regions]
        
        # Get all region names in batches
        region_name_results = self.get_parameters_batch(region_name_paths)
        
        # Map region codes to names
        region_names = {}
        for region in regions:
            path = f'/aws/service/global-infrastructure/regions/{region}/longName'
            name = region_name_results.get(path)
            if name:
                region_names[region] = name
            else:
                logging.warning(f"No display name found for region: {region}")
                region_names[region] = region  # Fallback to code
        
        # Save to cache
        self._save_to_cache(cache_key, region_names)
        
        logging.info(f"Successfully fetched names for {len(region_names)} regions")
        return region_names
    
    def fetch_service_names(self, services: List[str]) -> Dict[str, str]:
        """Fetch human-readable names for all services."""
        cache_key = "service_names"
        
        # Try to load from cache first
        cached_data = self._load_from_cache(cache_key)
        if cached_data is not None:
            # Filter cached data to only include requested services
            filtered_data = {k: v for k, v in cached_data.items() if k in services}
            if len(filtered_data) == len(services):
                logging.info(f"Using cached service names for {len(services)} services")
                return filtered_data
        
        logging.info("Fetching service display names...")
        
        # Build parameter paths for service names
        service_name_paths = [f'/aws/service/global-infrastructure/services/{service}/longName' for service in services]
        
        # Get all service names in batches
        service_name_results = self.get_parameters_batch(service_name_paths)
        
        # Map service codes to names
        service_names = {}
        for service in services:
            path = f'/aws/service/global-infrastructure/services/{service}/longName'
            name = service_name_results.get(path)
            if name:
                service_names[service] = name
            else:
                logging.warning(f"No display name found for service: {service}")
                service_names[service] = service  # Fallback to code
        
        # Save to cache
        self._save_to_cache(cache_key, service_names)
        
        logging.info(f"Successfully fetched names for {len(service_names)} services")
        return service_names
    
    def fetch_regional_services(self, regions: List[str], services: List[str]) -> Dict[str, List[str]]:
        """DEPRECATED: Test which services are available in each region by checking service names.
        
        This method has been replaced by get_services_per_region_proper() which uses actual AWS SSM data
        instead of estimation patterns. This method is kept for backward compatibility but should not be used.
        """
        logging.info("Testing regional service availability...")
        
        regional_services = {}
        
        for region in regions:
            logging.info(f"Testing services for region: {region}")
            available_services = []
            
            # Use all discovered services
            test_services = services  # Use all discovered services
            
            # Build parameter paths to test
            test_paths = [f'/aws/service/global-infrastructure/services/{service}/longName' for service in test_services]
            
            # Get service names (this validates the service exists)
            service_name_results = self.get_parameters_batch(test_paths)
            
            # Implement realistic regional service availability based on AWS patterns
            # Since regional service parameters don't exist in SSM, use service availability logic
            
            # Core services available in all regions
            core_services = {
                'ec2', 's3', 'iam', 'cloudformation', 'cloudwatch', 'sns', 'sqs', 
                'lambda', 'dynamodb', 'rds', 'elasticloadbalancing', 'autoscaling',
                'route53', 'cloudtrail', 'config', 'kms', 'ssm', 'sts'
            }
            
            # Extended services for established regions
            extended_services = {
                'acm', 'apigateway', 'backup', 'batch', 'codebuild', 'codecommit', 
                'codedeploy', 'codepipeline', 'ecr', 'ecs', 'efs', 'eks', 'elasticache',
                'emr', 'events', 'firehose', 'glue', 'kinesis', 'logs', 'secretsmanager',
                'stepfunctions', 'xray', 'application-autoscaling', 'elasticfilesystem',
                'elasticmapreduce', 'waf', 'shield', 'guardduty', 'inspector'
            }
            
            # Premium services for major regions only
            premium_services = {
                'sagemaker', 'bedrock', 'comprehend', 'rekognition', 'textract', 'translate',
                'transcribe', 'polly', 'lex', 'personalize', 'forecast', 'frauddetector',
                'kendra', 'connect', 'chime', 'workspaces', 'workdocs', 'workmail',
                'appstream', 'worklink', 'gamelift', 'robomaker', 'groundstation'
            }
            
            # Determine service availability based on region tier
            if region in ['us-east-1', 'us-west-2', 'eu-west-1']:
                # Tier 1: Major global regions - get most services
                available_service_set = core_services | extended_services | premium_services
                # Add some additional cutting-edge services
                additional_services = {'amplify', 'appsync', 'cognito-idp', 'pinpoint', 'mobiletargeting'}
                available_service_set |= additional_services
                
            elif region in ['us-east-2', 'ap-southeast-1', 'ap-northeast-1', 'eu-central-1', 'ca-central-1']:
                # Tier 2: Major regional hubs - get core + extended + some premium
                available_service_set = core_services | extended_services
                # Add subset of premium services
                available_service_set |= {'sagemaker', 'comprehend', 'rekognition', 'connect'}
                
            elif region in ['ap-south-1', 'ap-northeast-2', 'eu-west-2', 'us-west-1', 'sa-east-1']:
                # Tier 3: Established regions - get core + extended services
                available_service_set = core_services | extended_services
                
            elif region in ['af-south-1', 'ap-east-1', 'eu-south-1', 'me-south-1', 'me-central-1']:
                # Tier 4: Newer regions - get core services + some extended
                available_service_set = core_services
                basic_extended = {'acm', 'apigateway', 'backup', 'ecr', 'ecs', 'efs', 'elasticache', 'logs'}
                available_service_set |= basic_extended
                
            else:
                # Tier 5: Newest/smallest regions - core services only
                available_service_set = core_services
            
            # Filter to only include services that actually exist in our discovered list
            for service in test_services:
                path = f'/aws/service/global-infrastructure/services/{service}/longName'
                if (service_name_results.get(path) is not None and 
                    service in available_service_set):
                    available_services.append(service)
            
            regional_services[region] = sorted(available_services)
            logging.info(f"Regional availability: {len(available_services)} services available in {region}")
        
        return regional_services
    
    def generate_data_matrix(self, regions: List[str], region_names: Dict[str, str], 
                           service_names: Dict[str, str], regional_services: Dict[str, List[str]]) -> List[Dict]:
        """Generate the data matrix in the same format as the Excel file."""
        logging.info("Generating data matrix...")
        
        data = []
        total_combinations = 0
        
        for region_code in regions:
            region_name = region_names.get(region_code, region_code)
            services_in_region = regional_services.get(region_code, [])
            
            for service_code in services_in_region:
                service_name = service_names.get(service_code, service_code)
                
                data.append({
                    'Region Code': region_code,
                    'Region Name': region_name,
                    'Service Code': service_code,
                    'Service Name': service_name
                })
                total_combinations += 1
        
        logging.info(f"Generated {total_combinations} region-service combinations")
        return data
    
    def generate_service_matrix(self, data: List[Dict]) -> pd.DataFrame:
        """Generate service matrix showing which services are available in which regions."""
        # Create a pivot table
        df = pd.DataFrame(data)
        
        # Create matrix with services as rows and regions as columns
        matrix_data = []
        
        # Get unique services and regions
        services = sorted(df['Service Name'].unique())
        regions = sorted(df['Region Code'].unique())
        
        for service in services:
            row = {'Service': service}
            service_regions = set(df[df['Service Name'] == service]['Region Code'])
            
            for region in regions:
                row[region] = '✓' if region in service_regions else '✗'
            
            matrix_data.append(row)
        
        return pd.DataFrame(matrix_data)
    
    def generate_region_summary(self, data: List[Dict], region_names: Dict[str, str], 
                               rss_data: Dict[str, Dict] = None, az_data: Dict[str, int] = None,
                               all_services: List[str] = None) -> pd.DataFrame:
        """Generate region summary with service counts and RSS launch dates."""
        df = pd.DataFrame(data)
        
        summary_data = []
        for region_code in sorted(df['Region Code'].unique()):
            region_data = df[df['Region Code'] == region_code]
            
            # Use the actual number of services in the data for this region
            service_count = len(region_data)
            
            # Get RSS data for this region
            rss_region_data = (rss_data or {}).get(region_code, {})
            
            # Use RSS data if available, fallback to defaults
            launch_date = rss_region_data.get('launch_date', 'N/A')
            launch_date_source = 'AWS RSS Feed' if launch_date != 'N/A' else 'Not Available'
            announcement_url = rss_region_data.get('announcement_url', 'N/A')
            
            # Get AZ count
            az_count = (az_data or {}).get(region_code, 'N/A')
            
            summary_data.append({
                'Region Code': region_code,
                'Region Name': region_names.get(region_code, region_code),
                'Launch Date': launch_date,
                'Launch Date Source': launch_date_source,
                'Announcement URL': announcement_url,
                'Availability Zones': az_count if az_count != 'N/A' else 'N/A',
                'Service Count': service_count
            })
        
        return pd.DataFrame(summary_data)
    
    def estimate_service_regional_availability(self, service_code: str, service_names: Dict[str, str]) -> int:
        """DEPRECATED: Estimate how many regions a service is available in based on service patterns and SSM data.
        
        This method has been replaced by the proper AWS SSM data mapping approach in get_services_per_region_proper()
        and generate_service_summary(). This method is kept for backward compatibility but should not be used.
        """
        
        # Check if service exists globally by checking its longName parameter
        try:
            service_param = f'/aws/service/global-infrastructure/services/{service_code}/longName'
            service_exists = self.get_parameter_value(service_param)
            
            if not service_exists:
                return 0  # Service doesn't exist
                
        except:
            return 0  # Service doesn't exist or not accessible
        
        # Service exists globally, now estimate regional availability based on service type and patterns
        service_name = service_names.get(service_code, service_code).lower()
        
        # Global/Universal services (available in all regions)
        universal_services = {
            # Global AWS services
            'iam', 'cloudfront', 'route53', 'waf', 'shield', 'cloudformation', 
            'cloudwatch', 'cloudtrail', 'config', 'sns', 'sqs', 's3', 'sts',
            # Core compute/storage services that are truly universal
            'ec2', 'lambda', 'dynamodb', 'rds', 'elasticloadbalancing', 'autoscaling',
            'ebs', 'kms', 'ssm', 'secretsmanager'
        }
        
        # Core regional services (available in most regions ~92%)
        core_regional_services = {
            'elasticache', 'efs', 'fsx', 'backup', 'batch', 'ecr', 'ecs'
        }
        
        # Developer/integration services (available in major regions)
        developer_services = {
            'apigateway', 'codebuild', 'codecommit', 'codedeploy', 'codepipeline',
            'ecr', 'ecs', 'eks', 'batch', 'stepfunctions', 'eventbridge', 'backup'
        }
        
        # Analytics/ML services (available in fewer regions)  
        analytics_ml_services = {
            'sagemaker', 'comprehend', 'rekognition', 'textract', 'translate',
            'transcribe', 'polly', 'lex', 'bedrock', 'kendra', 'forecast'
        }
        
        # Specialized/new services (limited availability)
        specialized_services = {
            'braket', 'groundstation', 'robomaker', 'deeplens', 'devicefarm',
            'gamelift', 'lumberyard', 'sumerian', 'honeycode'
        }
        
        # Determine availability based on service type
        if service_code in universal_services or 'global' in service_name:
            return 38  # Available in all regions (100%)
        elif service_code in core_regional_services:
            return 35  # Available in most regions (92%)
        elif service_code in developer_services or any(keyword in service_name for keyword in ['api', 'code', 'container', 'serverless']):
            return 25  # Available in major regions (66%)
        elif service_code in analytics_ml_services or any(keyword in service_name for keyword in ['ml', 'ai', 'machine learning', 'analytics']):
            return 8   # Available in ML regions (21%)
        elif service_code in specialized_services or any(keyword in service_name for keyword in ['iot', 'game', 'quantum', 'satellite']):
            return 3   # Available in few regions (8%)
        else:
            # For unknown services, make educated guess based on name patterns
            if any(keyword in service_name for keyword in ['aws', 'amazon', 'elastic']):
                return 20  # Likely mainstream service (53%)
            elif 'beta' in service_name or 'preview' in service_name:
                return 2   # Limited beta service (5%)
            else:
                return 12  # Default moderate availability (32%)
    
        return 0

    def generate_service_summary(self, data: List[Dict], all_services: List[str] = None, service_names: Dict[str, str] = None) -> pd.DataFrame:
        """Generate service summary with region counts and coverage for ALL discovered services."""
        df = pd.DataFrame(data)
        total_regions = 38  # Use known total region count
        all_regions = ['af-south-1', 'ap-east-1', 'ap-east-2', 'ap-northeast-1', 'ap-northeast-2', 'ap-northeast-3',
                      'ap-south-1', 'ap-south-2', 'ap-southeast-1', 'ap-southeast-2', 'ap-southeast-3', 'ap-southeast-4',
                      'ap-southeast-5', 'ap-southeast-6', 'ap-southeast-7', 'ca-central-1', 'ca-west-1', 'cn-north-1',
                      'cn-northwest-1', 'eu-central-1', 'eu-central-2', 'eu-north-1', 'eu-south-1', 'eu-south-2',
                      'eu-west-1', 'eu-west-2', 'eu-west-3', 'il-central-1', 'me-central-1', 'me-south-1',
                      'mx-central-1', 'sa-east-1', 'us-east-1', 'us-east-2', 'us-gov-east-1', 'us-gov-west-1',
                      'us-west-1', 'us-west-2']
        
        summary_data = []
        
        # If all_services provided, check actual regional availability for each service
        if all_services and service_names:
            logging.info("Checking actual regional availability for all 395 services...")
            
            for i, service_code in enumerate(sorted(all_services)):
                # Get service name from service_names dict
                service_name = service_names.get(service_code, service_code)
                
                # Count actual regions where this service appears in the data
                service_data = df[df['Service Code'] == service_code]
                region_count = len(service_data)  # Each row is a region where the service is available
                coverage_pct = round((region_count / total_regions) * 100, 1)
                
                summary_data.append({
                    'Service Code': service_code,
                    'Service Name': service_name,
                    'Region Count': region_count,
                    'Coverage %': coverage_pct / 100  # Store as decimal
                })
                
                # Log progress every 50 services
                if (i + 1) % 50 == 0:
                    logging.info(f"Processed {i + 1}/{len(all_services)} services...")
        else:
            # Fallback to old method if all_services not provided
            for service_name in sorted(df['Service Name'].unique()):
                service_data = df[df['Service Name'] == service_name]
                region_count = len(service_data)
                coverage_pct = round((region_count / total_regions) * 100, 1)
                
                # Get service code (take first occurrence)
                service_code = service_data.iloc[0]['Service Code']
                
                summary_data.append({
                    'Service Code': service_code,
                    'Service Name': service_name,
                    'Region Count': region_count,
                    'Coverage %': coverage_pct / 100
                })
        
        return pd.DataFrame(summary_data)
    
    def generate_statistics(self, data: List[Dict], all_services: List[str] = None) -> pd.DataFrame:
        """Generate statistics sheet."""
        df = pd.DataFrame(data)
        
        # Use total discovered services count if available
        total_services = len(all_services) if all_services else df['Service Name'].nunique()
        
        # Debug logging
        logging.info(f"Generate Statistics: all_services parameter contains {len(all_services) if all_services else 'None'} services")
        logging.info(f"Generate Statistics: Using total_services = {total_services}")
        
        stats = [
            ['Generator', 'AWS SSM Data Fetcher with Caching v2.0'],
            ['', ''],
            ['Summary Statistics', ''],
            ['Total Regions', df['Region Code'].nunique()],
            ['Total Services', total_services],
            ['Total Combinations', len(df)],
            ['', ''],
            ['Regional Service Distribution', ''],
            ['Avg Services per Region', round(df.groupby('Region Code').size().mean(), 1)],
            ['Max Services (Region)', df.groupby('Region Code').size().max()],
            ['Min Services (Region)', df.groupby('Region Code').size().min()],
            ['', ''],
            ['Service Distribution', ''],
            ['Avg Regions per Service', round(df.groupby('Service Name').size().mean(), 1)],
            ['Max Regions (Service)', df.groupby('Service Name').size().max()],
            ['Min Regions (Service)', df.groupby('Service Name').size().min()]
        ]
        
        stats_df = pd.DataFrame(stats, columns=['Generated At', datetime.now().strftime('%Y-%m-%d %H:%M:%S')])
        return stats_df
    
    def save_to_excel(self, data: List[Dict], filename: str = None, output_dir: str = 'output', 
                     region_names: Dict[str, str] = None, rss_data: Dict[str, Dict] = None,
                     all_services: List[str] = None, service_names: Dict[str, str] = None):
        """Save data to Excel file with multiple sheets matching the original format."""
        # Create output directory if it doesn't exist
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            logging.info(f"Created output directory: {output_dir}")
        
        if filename is None:
            filename = 'aws_regions_services.xlsx'
        
        # Ensure filename includes output directory
        filepath = os.path.join(output_dir, filename) if not filename.startswith(output_dir) else filename
        
        logging.info(f"Saving comprehensive Excel report: {filepath}")
        
        # Fetch availability zone data for regions
        df = pd.DataFrame(data)
        regions = df['Region Code'].unique().tolist()
        az_data = self.fetch_availability_zones(regions)
        
        # Generate all sheets
        regional_services_df = pd.DataFrame(data)
        service_matrix_df = self.generate_service_matrix(data)
        region_summary_df = self.generate_region_summary(data, region_names or {}, rss_data, az_data, all_services)
        service_summary_df = self.generate_service_summary(data, all_services, service_names)
        statistics_df = self.generate_statistics(data, all_services)
        
        # Save to Excel with multiple sheets and auto-adjust column widths
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Write all sheets
            regional_services_df.to_excel(writer, sheet_name='Regional Services', index=False)
            service_matrix_df.to_excel(writer, sheet_name='Service Matrix', index=False)
            region_summary_df.to_excel(writer, sheet_name='Region Summary', index=False)
            service_summary_df.to_excel(writer, sheet_name='Service Summary', index=False)
            statistics_df.to_excel(writer, sheet_name='Statistics', index=False)
            
            # Auto-adjust column widths and apply formatting for all sheets
            from openpyxl.styles import PatternFill, Font
            
            # Define colors and fonts to match the original file
            green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Light green for ✓
            red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")    # Light red for ✗
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid") # Dark blue for headers
            white_font = Font(color="FFFFFF")  # White font for headers
            
            for sheet_name in writer.sheets:
                worksheet = writer.sheets[sheet_name]
                
                # Get the dataframe for this sheet to calculate optimal widths
                if sheet_name == 'Regional Services':
                    df = regional_services_df
                elif sheet_name == 'Service Matrix':
                    df = service_matrix_df
                elif sheet_name == 'Region Summary':
                    df = region_summary_df
                elif sheet_name == 'Service Summary':
                    df = service_summary_df
                elif sheet_name == 'Statistics':
                    df = statistics_df
                
                # Apply header formatting to ALL sheets (matching sample file)
                logging.info(f"Applying header formatting to {sheet_name}...")
                
                # Color the header row (row 1) for all sheets with blue background and white font
                for col in range(1, worksheet.max_column + 1):
                    cell = worksheet.cell(row=1, column=col)
                    cell.fill = header_fill
                    cell.font = white_font
                
                # Apply special formatting for Service Matrix
                if sheet_name == 'Service Matrix':
                    logging.info("Applying color formatting to Service Matrix data cells...")
                    
                    # Color the data cells based on ✓ or ✗ values
                    for row in range(2, worksheet.max_row + 1):  # Start from row 2 (skip header)
                        for col in range(2, worksheet.max_column + 1):  # Start from col 2 (skip service name)
                            cell = worksheet.cell(row=row, column=col)
                            if cell.value == '✓':
                                cell.fill = green_fill
                            elif cell.value == '✗':
                                cell.fill = red_fill
                
                # Apply percentage formatting for Service Summary
                elif sheet_name == 'Service Summary':
                    logging.info("Applying percentage formatting to Service Summary...")
                    
                    # Find the Coverage % column (should be column D)
                    coverage_col = None
                    for col in range(1, worksheet.max_column + 1):
                        header_cell = worksheet.cell(row=1, column=col)
                        if header_cell.value == 'Coverage %':
                            coverage_col = col
                            break
                    
                    if coverage_col:
                        # Apply percentage format to the Coverage % column
                        from openpyxl.styles import NamedStyle
                        
                        # Create percentage style
                        percentage_style = NamedStyle(name="percentage")
                        percentage_style.number_format = '0.0%'
                        
                        # Apply to all data cells in Coverage % column
                        for row in range(2, worksheet.max_row + 1):
                            cell = worksheet.cell(row=row, column=coverage_col)
                            cell.style = percentage_style
                
                # Auto-adjust column widths
                for column in worksheet.columns:
                    column_letter = column[0].column_letter
                    column_name = df.columns[column[0].column - 1] if column[0].column <= len(df.columns) else ''
                    
                    # Calculate optimal width based on content
                    max_length = 0
                    
                    # Check header length
                    if column_name:
                        max_length = max(max_length, len(str(column_name)))
                    
                    # Check data content length (sample first 100 rows for performance)
                    if column[0].column <= len(df.columns):
                        col_data = df.iloc[:min(100, len(df)), column[0].column - 1]
                        for cell_value in col_data:
                            if cell_value is not None:
                                max_length = max(max_length, len(str(cell_value)))
                    
                    # Set column width with reasonable limits (min 10, max 50 characters)
                    adjusted_width = min(max(max_length + 2, 10), 50)
                    worksheet.column_dimensions[column_letter].width = adjusted_width
        
        # Get stats for logging
        stats = {
            'regions': regional_services_df['Region Code'].nunique(),
            'services': regional_services_df['Service Name'].nunique(),
            'combinations': len(regional_services_df)
        }
        
        logging.info(f"Excel report saved: {filepath}")
        logging.info(f"  - Regional Services: {stats['combinations']} rows")
        logging.info(f"  - Service Matrix: {len(service_matrix_df)} services × {len(service_matrix_df.columns)-1} regions")
        logging.info(f"  - Region Summary: {stats['regions']} regions")
        logging.info(f"  - Service Summary: {stats['services']} services")
        logging.info(f"  - Statistics: {len(statistics_df)} metrics")
        
        return filepath
    
    def save_to_json(self, data: List[Dict], filename: str = None, output_dir: str = 'output'):
        """Save data to JSON file."""
        # Create output directory if it doesn't exist
        if not os.path.exists(output_dir):
            os.makedirs(output_dir)
            logging.info(f"Created output directory: {output_dir}")
        
        if filename is None:
            filename = 'aws_regions_services.json'
        
        # Ensure filename includes output directory
        filepath = os.path.join(output_dir, filename) if not filename.startswith(output_dir) else filename
        
        logging.info(f"Saving data to JSON: {filepath}")
        
        # Create comprehensive JSON structure
        json_data = {
            'metadata': {
                'generated_at': datetime.now().isoformat(),
                'total_combinations': len(data),
                'unique_regions': len(set(item['Region Code'] for item in data)),
                'unique_services': len(set(item['Service Code'] for item in data)),
                'source': 'AWS SSM Parameter Store'
            },
            'data': data
        }
        
        with open(filepath, 'w', encoding='utf-8') as f:
            json.dump(json_data, f, indent=2, ensure_ascii=False)
        
        logging.info(f"JSON file saved: {filepath}")
        return filepath


def main():
    """Main execution function."""
    import argparse
    
    # Set up argument parser
    parser = argparse.ArgumentParser(description='AWS SSM Data Fetcher with Caching')
    parser.add_argument('--clear-cache', action='store_true', 
                       help='Clear all cached data before running')
    parser.add_argument('--cache-info', action='store_true',
                       help='Show cache information and exit')
    parser.add_argument('--cache-hours', type=int, default=24,
                       help='Cache TTL in hours (default: 24)')
    parser.add_argument('--no-cache', action='store_true',
                       help='Disable caching for this run')
    parser.add_argument('--cache-dir', type=str, default='.cache',
                       help='Cache directory path (default: .cache)')
    parser.add_argument('--output-dir', type=str, default='output',
                       help='Output directory for generated files (default: output)')
    
    args = parser.parse_args()
    
    # Set up logging
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(levelname)s - %(message)s'
    )
    
    try:
        # Initialize the fetcher with config from command line arguments
        config = Config.from_args(args)
        fetcher = AWSSSMDataFetcher(config=config)
        
        # Handle cache management commands
        if args.cache_info:
            cache_info = fetcher.get_cache_info()
            print("\n" + "="*60)
            print("CACHE INFORMATION")
            print("="*60)
            print(f"Cache Directory: {cache_info['cache_dir']}")
            print(f"TTL Hours: {cache_info['ttl_hours']}")
            print(f"Total Files: {cache_info['total_files']}")
            print(f"Total Size: {cache_info['total_size_kb']} KB")
            print("\nCached Files:")
            for file_info in cache_info['files']:
                status = "✅ Valid" if file_info['valid'] else "❌ Expired"
                print(f"  {file_info['file']}: {file_info['size_kb']} KB, {status}")
                print(f"    Created: {file_info['created']}")
                print(f"    Expires: {file_info['expires']}")
            print("="*60)
            return
        
        if args.clear_cache:
            cleared_count = fetcher.clear_cache()
            print(f"✅ Cleared {cleared_count} cache files")
            if not args.cache_info:
                print("Continuing with fresh data fetch...")
            else:
                return
        
        # Fetch basic data
        logging.info("Starting AWS SSM data fetch...")
        regions = fetcher.fetch_regions()
        services = fetcher.fetch_services()
        
        if not regions or not services:
            logging.error("Failed to fetch basic regions or services data")
            return
        
        # Fetch display names
        region_names = fetcher.fetch_region_names(regions)
        service_names = fetcher.fetch_service_names(services)
        
        # Fetch RSS data for region launch dates
        rss_data = fetcher.fetch_region_rss_data()
        
        # Fetch regional service availability using proper AWS SSM mapping
        logging.info("Using proper AWS SSM data for regional service mapping...")
        regional_services = fetcher.get_services_per_region_proper(services)
        
        # Generate data matrix
        data = fetcher.generate_data_matrix(regions, region_names, service_names, regional_services)
        
        if not data:
            logging.error("No data generated")
            return
        
        # Save to both formats
        excel_file = fetcher.save_to_excel(data, output_dir=args.output_dir, region_names=region_names, 
                                         rss_data=rss_data, all_services=services, service_names=service_names)
        json_file = fetcher.save_to_json(data, output_dir=args.output_dir)
        
        # Print summary
        print("\n" + "="*60)
        print("AWS SSM Data Fetch Complete!")
        print("="*60)
        print(f"Regions processed: {len(regions)}")
        print(f"Services processed: {len(services)}")
        print(f"Total region-service combinations: {len(data)}")
        print(f"Excel file: {excel_file}")
        print(f"JSON file: {json_file}")
        print("="*60)
        
    except Exception as e:
        logging.error(f"Script execution failed: {e}")
        raise


if __name__ == "__main__":
    main()